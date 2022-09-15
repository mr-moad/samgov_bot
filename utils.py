import os
import time

import pandas as pd
from openpyxl import load_workbook
import traceback
def save_to_excel(data):
    sheet_name = data.get("keyword") or data.get("naics") or "unsorted"
    try:
        data = [data]
        if not os.path.isfile("records.xlsx"):
            dataframe = pd.DataFrame(data)
            data_to_excel = pd.ExcelWriter('records.xlsx', engine='xlsxwriter')
            dataframe.to_excel(data_to_excel, sheet_name=sheet_name,index=False)
            data_to_excel.save()
            data_to_excel.close()
        else:
            wb = load_workbook("records.xlsx", read_only=True)

            if str(sheet_name) in wb.sheetnames:
                file = pd.read_excel('records.xlsx',sheet_name=str(sheet_name))
                new_df = file.append(data)
                with pd.ExcelWriter('records.xlsx', engine='openpyxl', mode="a", if_sheet_exists="overlay") as writes:
                        new_df.to_excel(writes,str(sheet_name),index=False)
            else:
                dataframe = pd.DataFrame(data)
                with pd.ExcelWriter('records.xlsx', engine='openpyxl', mode="a", if_sheet_exists="overlay") as writes:
                    dataframe.to_excel(writes, str(sheet_name), index=False)
    except:
        print(traceback.format_exc())